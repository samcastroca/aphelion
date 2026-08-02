"""Catálogo canónico de documentos, derivado del inventario oficial de ADL.

El identificador de cada documento no se inventa: se replica la fórmula que el
Excel define en la columna DOC_ID,

    = fenomeno & "-" & codigo_observatorio & "-" & TEXT(consecutivo, "000")

donde el consecutivo cuenta las apariciones previas del par (fenómeno, código)
en el orden de filas de la hoja. El resultado es `F1-CSET-001`.

El campo `fuente` se toma literal del inventario porque es la clave con la que
el jurado empareja los documentos contra el ground truth (especificación §10.2.1).
"""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass, asdict
from pathlib import Path

import openpyxl

from . import config


@dataclass(frozen=True)
class Documento:
    doc_id: str
    fuente: str  # nombre estandarizado, clave de emparejamiento
    fenomeno: int  # 1, 2 o 3
    observatorio: str
    codigo: str
    tipo: str  # PDF, JSON, CSV, Excel, Imagen, Texto, Otro
    ruta_rel: str  # ruta relativa dentro del corpus, para trazabilidad

    @property
    def ruta(self) -> Path:
        return config.CORPUS / self.ruta_rel

    @property
    def formato(self) -> str:
        """Extensión real del archivo, en minúsculas y sin punto."""
        return Path(self.fuente).suffix.lstrip(".").lower()

    def to_dict(self) -> dict:
        return asdict(self)


def _codigos_por_observatorio(wb) -> dict[str, str]:
    """Mapa observatorio -> código, leído de la hoja 'Indice'."""
    codigos: dict[str, str] = {}
    filas = wb["Indice"].iter_rows(values_only=True)
    next(filas, None)  # cabecera
    for fila in filas:
        if not fila or not fila[1] or not fila[2]:
            continue
        observatorio, codigo = str(fila[1]), str(fila[2])
        if codigo.startswith("="):  # fórmula sin evaluar
            continue
        codigos[observatorio] = codigo
    return codigos


def cargar(inventario: Path | None = None) -> list[Documento]:
    """Devuelve los documentos del inventario, en el orden de la hoja."""
    ruta = inventario or config.INVENTARIO
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=False)

    codigos = _codigos_por_observatorio(wb)

    filas = wb["Inventario de Archivos"].iter_rows(values_only=True)
    next(filas, None)  # cabecera

    documentos: list[Documento] = []
    contador: Counter[tuple[str, str]] = Counter()

    for fila in filas:
        if not fila or not fila[0]:
            continue
        fen_txt, observatorio, _, _, nombre, carpeta, tipo = fila[:7]
        if not nombre:
            continue

        codigo = codigos.get(str(observatorio))
        if codigo is None:
            raise KeyError(f"observatorio sin código en la hoja Indice: {observatorio!r}")

        clave = (str(fen_txt), codigo)
        contador[clave] += 1

        documentos.append(
            Documento(
                doc_id=f"{fen_txt}-{codigo}-{contador[clave]:03d}",
                fuente=str(nombre),
                fenomeno=int(str(fen_txt).lstrip("F")),
                observatorio=str(observatorio),
                codigo=codigo,
                tipo=str(tipo),
                ruta_rel=str(Path(str(carpeta)) / str(nombre)),
            )
        )

    return documentos


def verificar(documentos: list[Documento]) -> dict:
    """Contrasta el catálogo contra el estado real del disco."""
    faltan = [d for d in documentos if not d.ruta.exists()]
    ids = Counter(d.doc_id for d in documentos)
    fuentes = Counter(d.fuente for d in documentos)

    return {
        "total": len(documentos),
        "en_disco": len(documentos) - len(faltan),
        "faltantes": [d.ruta_rel for d in faltan],
        "doc_id_duplicados": [k for k, v in ids.items() if v > 1],
        "fuentes_duplicadas": sum(1 for v in fuentes.values() if v > 1),
        "por_tipo": dict(Counter(d.tipo for d in documentos)),
        "por_fenomeno": dict(Counter(d.fenomeno for d in documentos)),
    }
